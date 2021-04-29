from openpyxl import Workbook
from administration.models import *
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date, timedelta, datetime



@login_required
def sales_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	sales = kwargs.get('sales')

	file_name = f"{business.name}_Sales_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Sales Report'

	 # Define the titles for row
	row = ['#','ORDER NO.','PRODUCT','QUANTITY','TOTAL','PROFIT','PAID','DISCOUNT','TAX', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for sale in sales:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			sale.order_no,
			sale.inventory.product.name,
			sale.quantity,
			sale.total,
			sale.profit,
			sale.amount_paid,
			f'{sale.discount} {sale.discount_unit}',
			f'{sale.tax} {sale.tax_unit}',
			sale.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response  		


@login_required
def inventory_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	inventories = kwargs.get('inventories')

	file_name = f"{business.name}_Inventory_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Inventory Report'

	 # Define the titles for row
	row = ['#','PRODUCT','QUANTITY','AVAILABLE','DAMAGE','PRODUCT COST','SELL PRICE','COGS', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for inventory in inventories:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			inventory.product.name,
			inventory.quantity,
			inventory.remain-inventory.damage,
			inventory.damage,			
			inventory.product_cost,
			inventory.product.sell_price,
			inventory.cogs,
			inventory.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def procurement_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	local_purchases = kwargs.get('local_purchases')
	purchases = kwargs.get('purchases') 

	file_name = f"{business.name}_Procurement_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Local Purchases Report'
	ws1.sheet_properties.tabColor = "007bff"

	 # Define the titles for row
	row = ['#','LPO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in local_purchases:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			local.lpo_no,
			local.supplier.name,
			local.delivery.location,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value

	# Purchases order
	ws2 = workbook.create_sheet()
	ws2.title = f'{business.name} Purchases Report'
	ws2.sheet_properties.tabColor = "dc3545"

	 # Define the titles for row
	row = ['#','PO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = ws2.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in purchases:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			local.po_no,
			local.supplier.name,
			local.delivery.location,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws2.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def opex_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	expenses = kwargs.get('expense_objects')

	file_name = f"{business.name}_OPEX_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} OPEX Report'

	 # Define the titles for columns
	columns = ['#','NAME','CATEGORY','COST','DATE',]
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for expense in expenses:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			expense.name,
			expense.category,
			expense.cost,			
			expense.date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def payroll_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	takehome = kwargs.get('takehome')

	file_name = f"{business.name}_Payroll_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Payroll Report'

	 # Define the titles for columns
	columns = ['#','EMPLOYEE ID','NAME','POSITION','DEPARTMENT','SALARY','TAKEHOME','PAYE','NSSF','HESLB','WCF','BONUS','OVERTIME','DEDUCTION','SDL','CREATED']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	for worker in takehome:  	
		# Define the data for each cell in the row 
		loan = Loan.objects.filter(employee=worker.payroll.employee, business=business).first()

		if worker.payroll.paye:			
				if 'nssf' in worker.payroll.tax_rate:
						nssf_funds = worker.payroll.employee.salary	* 0.1
				else:
					nssf_funds = 0
				if 'wcf' in worker.payroll.tax_rate:
							wcf_funds = worker.payroll.employee.salary	* 0.01
				else:
					wcf_funds = 0
				if 'loan board' in worker.payroll.tax_rate:
							loan_funds = worker.payroll.employee.salary	* 0.15
				else:
					loan_funds = 0
		salary = worker.salary + nssf_funds + wcf_funds + loan_funds

	
		if loan:
			if worker.payroll.employee.employee_loan.debt == True:
					paye = worker.payroll.employee.salary - salary + worker.payroll.overtime + worker.payroll.bonus - worker.payroll.deduction - worker.payroll.sdl_amount - worker.payroll.employee.employee_loan.amount_paid
		else: 
			paye = worker.payroll.employee.salary - salary + worker.payroll.overtime + worker.payroll.bonus - worker.payroll.deduction - worker.payroll.sdl_amount

		if 'nssf' in worker.payroll.tax_rate:
					nssf = worker.payroll.employee.salary	* 0.1
		else:
			nssf = 0
		if 'wcf' in worker.payroll.tax_rate:
					wcf = worker.payroll.employee.salary	* 0.01
		else:
			wcf = 0
		if 'loan board' in worker.payroll.tax_rate:
					loan = worker.payroll.employee.salary	* 0.15
		else:
			loan = 0

		row = [
			row_num,
			worker.payroll.employee.id_no,
			worker.payroll.employee.full_name,
			worker.payroll.employee.position,
			worker.payroll.employee.department.name,
			worker.payroll.employee.salary,
			worker.salary,	
			paye,
			nssf,
			loan,
			wcf,
			worker.payroll.bonus,
			worker.payroll.overtime,
			worker.payroll.deduction,
			worker.payroll.sdl_amount,	
			worker.payroll.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 
	



@login_required
def trial_balance_export_excel(request, **kwargs):
	id = kwargs.get('id')
	business = Business.objects.filter(id=id).first()
	fr = request.GET.get('fr')
	to = request.GET.get('to')	
	business = Business.objects.filter(id=id).first()
	nssf = Payroll.objects.filter(tax_rate__icontains='nssf', business=business, created_at__date__gte=fr, created_at__date__lte=to)
	wcf = Payroll.objects.filter(tax_rate__icontains='wcf', business=business, created_at__date__gte=fr, created_at__date__lte=to)
	loan_board = Payroll.objects.filter(tax_rate__icontains='loan board', business=business, created_at__date__gte=fr, created_at__date__lte=to)
	paye_objs = Payroll.objects.filter(paye=True, business=business, created_at__date__gte=fr, created_at__date__lte=to)
	takehome_total = Takehome.objects.filter(business=business).aggregate(Sum('salary'))['salary__sum']
	sdl = Payroll.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('sdl_amount'))['sdl_amount__sum']   		
	taxes = Tax.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('remain'))['remain__sum']
	interest_remaining = Interest.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('remaining'))['remaining__sum']
	interest_repayment = Interest.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('repayment'))['repayment__sum']
	sales = Sale.objects.filter(branch__business=business, status='Completed', created_at__date__gte=fr, created_at__date__lte=to).all().aggregate(total_paid=Sum('amount_paid'))
	inventory_qs = Inventory.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).annotate(available=F('remain')-F('damage'))			
	inventories = inventory_qs.filter(available__gt=0, ).order_by('-pk')
	expenses = Expense.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']
	liabilities = Liability.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']
	fixed_assets = AccountsFixedAsset.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('value'))['value__sum']
	current_assets = AccountsCurrentAsset.objects.filter(business=business, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']

	sales = sales['total_paid']



	cogs = 0
	for i in inventories:
		cogs = cogs + i.cogs

	if fixed_assets is None:
		fixed_assets = 0
	if current_assets is None:
		current_assets = 0
	if interest_remaining is None:
		interest_remaining = 0	
	if interest_repayment is None:
		interest_repayment = 0		
	
	interests = interest_repayment - interest_remaining		

	assets = fixed_assets + current_assets		


	paye_total = 0

	if paye_objs:     	

		for worker in paye_objs:

			salary = worker.employee.salary

			if salary <= 270000:

				paye_total += 0

			elif 270000 < salary <= 520000:

				tax_payment = (salary - 270000) * 0.09 
				paye_total += tax_payment


			elif 520000 < salary <= 760000:

				tax_payment = (salary - 520000) * 0.2 + 22500 
				paye_total += tax_payment

			elif 760000 < salary <= 1000000:

				tax_payment = (salary - 760000) * 0.25 + 70500 
				paye_total += tax_payment


			elif 1000000 < salary:

				tax_payment = (salary - 1000000) * 0.3 + 130500 
				paye_total += tax_payment          


	nssf_funds = 0
	wcf_funds = 0
	loan_board_funds = 0				

	total_funds = 0
	if nssf:
		# print(nssf)
		for worker in nssf:
			total_funds += worker.employee.salary
		nssf_funds = total_funds * 0.1 
		# print(f"NSSF : {int(nssf_funds)}")
		total_funds = 0

	if wcf:
		# print(wcf)
		for worker in wcf:
			total_funds += worker.employee.salary            	
		wcf_funds = total_funds * 0.01 
		# print(f"WCF : {int(wcf_funds)}")
		total_funds = 0

	if loan_board:
		# print(loan_board)
		for worker in loan_board:
			total_funds += worker.employee.salary
		loan_board_funds = total_funds * 0.15 
		# print(f"LOAN BOARD : {int(loan_board_funds)}")


	if takehome_total is None:
		takehome_total = 0
	if sdl is None:
		sdl = 0	
	if taxes is None:
		taxes = 0
	if interests is None:
		interests = 0					
	if liabilities is None:
		liabilities = 0
	if assets is None:
		assets = 0
	if cogs is None:
		cogs = 0
	if paye_total is None:
		paye_total = 0	
	if expenses is None:
		expenses = 0
	if sales is None:
		sales = 0



	total_debit = int(nssf_funds) + int(paye_total) + int(loan_board_funds) + int(wcf_funds) + int(takehome_total) + int(sdl) + int(interests) + int(taxes) + int(expenses) + int(cogs) + int(liabilities)
	total_credit = int(sales) + int(assets)	

	file_name = f"{business.name}_Trial_Balance_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Trial Balance Report'

	 # Define the titles for row
	row1 = ['DESCRIPTION', 'DEBIT', 'CREDIT']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row1, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	row3 = ['Assets', '0', int(assets)]
	row_num = 3

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row3, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row4 = ['Liabilities','0', int(liabilities)]
	row_num = 4

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row4, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row5 = ['Salaries', int(takehome_total), '0']
	row_num = 5

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row5, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row6 = ['Taxes', int(taxes), '0']
	row_num = 6

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row6, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row7 = ['Interests', int(interests), '0']
	row_num = 7

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row7, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row8 = ['Paye', int(paye_total), '0']
	row_num = 8

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row8, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row9 = ['Sdl', int(sdl), '0']
	row_num = 9

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row9, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row10 = ['Expenses', int(expenses), '0']
	row_num = 10

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row10, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row11 = ['Cogs', int(cogs), '0']
	row_num = 11

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row11, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row12 = ['Sales', '0', int(sales)]
	row_num = 12

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row12, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row13 = ['Heslb', int(loan_board_funds), '0']
	row_num = 13

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row13, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row14 = ['Nssf', int(nssf_funds), '0']
	row_num = 14

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row14, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row15 = ['Wcf', int(wcf_funds), '0']
	row_num = 15

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row15, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	row16 = ['Total', int(total_debit), int(total_credit)]
	row_num = 16

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row16, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title		



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response	


def balance_sheet_export_excel(request, **kwargs):
	business = kwargs.get('business')
	assets = kwargs.get('assets')
	liabilities = kwargs.get('liabilities')
	equity = kwargs.get('equity')

	file_name = f"{business.name}_Balance_Sheet_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Balance Sheet Report'

	 # Define the titles for columns


	columns = ['Assets','Liabilities','Equity']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	row = [assets, liabilities,equity]
	row_num = 2

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 

def income_statement_export_excel(request, **kwargs):
	summary = kwargs.get('summary')	
	business = kwargs.get('business')
	operating_revenue = summary['total_paid']
	operational_expense = summary['opex']
	tax_interest = summary['tax_interest']
	cogs = summary['cogs']
	total_income = summary['total_paid']
	net_income = summary['net_income']
	ebitda = summary['ebitda']
	ebit = summary['ebit']

	file_name = f"{business.name}_Income_Statement_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Income Statement'

	 # Define the titles for columns
	columns = ['PARTICULARS', 'AMOUNT']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	# Define the titles for columns
	columns = ['OPERATING REVENUE', operating_revenue]
	row_num = 2

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	# Define the titles for columns
	columns = ['TOTAL INCOME', total_income]
	row_num = 3

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['OPERATIONAL EXPENSE', operational_expense]
	row_num = 4

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['COGS', round(cogs,2)]
	row_num = 5

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['EBIT', ebit]
	row_num = 6

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['EBITDA', ebitda]
	row_num = 7

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	



	 # Define the titles for columns
	columns = ['TAX & INTEREST', tax_interest]
	row_num = 8

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['NET INCOME', net_income]
	row_num = 9

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response	

def cashbook_export_excel(request, **kwargs):
	business = kwargs.get('business')
	cashflow = kwargs.get('cashflow')

	file_name = f"{business.name}_OPEX_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} OPEX Report'

	 # Define the titles for columns
	columns = ['#','DATE CREATED','DESCRIPTION','DEBIT','CREDIT','BALANCE']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for cashbook in cashflow:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			cashbook.created_at.strftime('%d-%m-%Y'),
			cashbook.description,
			cashbook.debit,			
			cashbook.credit,
			cashbook.balance,
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response