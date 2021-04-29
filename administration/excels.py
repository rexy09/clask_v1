from openpyxl import Workbook
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date, timedelta, datetime
from business.models import Loan



@login_required
def sales_report_export_excel(request, **kwargs):
	sales = kwargs.get('sales')
	
	file_name = f"General_Sales_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'General Sales Report'

	 # Define the titles for columns
	columns = ['#','ORDER NO.','PRODUCT','QUANTITY','TOTAL','PROFIT','PAID','DISCOUNT','TAX', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for sale in sales:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			sale.order_no,
			sale.inventory.product.name,
			sale.quantity,
			sale.total,
			sale.profit,
			sale.amount_paid,
			f'{sale.discount} {sale.discount_unit}',
			f'{sale.tax} {sale.tax_unit}',
			sale.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response  	



@login_required
def inventory_report_export_excel(request, **kwargs):
	inventories = kwargs.get('inventories')

	file_name = f"General_Inventory_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'General Inventory Report'

	 # Define the titles for columns
	columns = ['#','PRODUCT','QUANTITY','AVAILABLE','DAMAGE','PRODUCT COST','SELL PRICE','COGS', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for inventory in inventories:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			inventory.product.name,
			inventory.quantity,
			inventory.remain-inventory.damage,
			inventory.damage,			
			inventory.product_cost,
			inventory.product.sell_price,
			inventory.cogs,
			inventory.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def procurement_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	local_purchases = kwargs.get('local_purchases')
	purchases = kwargs.get('purchases') 

	file_name = f"General_Procurement_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'General Local Purchases Report'
	ws1.sheet_properties.tabColor = "007bff"

	 # Define the titles for columns
	columns = ['#','LPO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in local_purchases:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			local.lpo_no,
			local.supplier.name,
			local.delivery.location,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value

	# Purchases order
	ws2 = workbook.create_sheet()
	ws2.title = f'General Purchases Report'
	ws2.sheet_properties.tabColor = "dc3545"

	 # Define the titles for columns
	columns = ['#','PO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = ws2.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in purchases:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			local.po_no,
			local.supplier.name,
			local.delivery.location,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws2.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def opex_report_export_excel(request, **kwargs):
	expenses = kwargs.get('expense_objects')

	file_name = f"General_OPEX_ Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'General OPEX Report'

	 # Define the titles for columns
	columns = ['#','NAME','CATEGORY','COST','DATE',]
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for expense in expenses:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			expense.name,
			expense.category,
			expense.cost,			
			expense.date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 	


@login_required
def payroll_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	takehome = kwargs.get('takehome')

	file_name = f"Payroll_General_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'Payroll General Report'

	 # Define the titles for columns
	columns = ['#','EMPLOYEE ID','NAME','POSITION','DEPARTMENT','SALARY','TAKEHOME','PAYE','NSSF','HESLB','WCF','BONUS','OVERTIME','DEDUCTION','SDL','CREATED']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	for worker in takehome:  	
		# Define the data for each cell in the row 
		loan = Loan.objects.filter(employee=worker.payroll.employee, business=business).first()

		if worker.payroll.paye:			
				if 'nssf' in worker.payroll.tax_rate:
						nssf_funds = worker.payroll.employee.salary	* 0.1
				else:
					nssf_funds = 0
				if 'wcf' in worker.payroll.tax_rate:
							wcf_funds = worker.payroll.employee.salary	* 0.01
				else:
					wcf_funds = 0
				if 'loan board' in worker.payroll.tax_rate:
							loan_funds = worker.payroll.employee.salary	* 0.15
				else:
					loan_funds = 0
		salary = worker.salary + nssf_funds + wcf_funds + loan_funds

	
		if loan:
			if worker.payroll.employee.employee_loan.debt == True:
					paye = worker.payroll.employee.salary - salary + worker.payroll.overtime + worker.payroll.bonus - worker.payroll.deduction - worker.payroll.sdl_amount - worker.payroll.employee.employee_loan.amount_paid
		else: 
			paye = worker.payroll.employee.salary - salary + worker.payroll.overtime + worker.payroll.bonus - worker.payroll.deduction - worker.payroll.sdl_amount

		if 'nssf' in worker.payroll.tax_rate:
					nssf = worker.payroll.employee.salary	* 0.1
		else:
			nssf = 0
		if 'wcf' in worker.payroll.tax_rate:
					wcf = worker.payroll.employee.salary	* 0.01
		else:
			wcf = 0
		if 'loan board' in worker.payroll.tax_rate:
					loan = worker.payroll.employee.salary	* 0.15
		else:
			loan = 0

		row = [
			row_num,
			worker.payroll.employee.id_no,
			worker.payroll.employee.full_name,
			worker.payroll.employee.position,
			worker.payroll.employee.department.name,
			worker.payroll.employee.salary,
			worker.salary,	
			paye,
			nssf,
			loan,
			wcf,
			worker.payroll.bonus,
			worker.payroll.overtime,
			worker.payroll.deduction,
			worker.payroll.sdl_amount,	
			worker.payroll.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 